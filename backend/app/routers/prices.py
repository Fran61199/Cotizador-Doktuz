"""API: descarga de plantilla XLSX, importación y listado/edición 1x1 de precios."""
import io
from typing import List, Dict, Any, Optional

from fastapi import APIRouter, UploadFile, File, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.models.db_models import Test, Clinic, Price
from app.services.price_import_service import import_prices_from_rows

try:
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None  # type: ignore

router = APIRouter()


class PriceUpdateBody(BaseModel):
    test_id: int
    clinic_id: Optional[int] = None  # None = Lima
    ingreso: float = 0
    periodico: float = 0
    retiro: float = 0


class AddPriceBody(BaseModel):
    """Añade una prueba (crea test si no existe) y su precio para la sede."""
    test_name: str
    category: str
    clinic_id: Optional[int] = None  # None = Lima
    ingreso: float = 0
    periodico: float = 0
    retiro: float = 0


class DeletePriceBody(BaseModel):
    """Elimina precio(s) de una prueba."""
    test_id: int
    scope: str  # "clinic" | "all_provincia" | "lima" | "all"
    clinic_id: Optional[int] = None  # Requerido si scope = "clinic"


class PriceRow(BaseModel):
    test_id: int
    test_name: str
    category: str
    price_id: Optional[int] = None
    ingreso: float = 0
    periodico: float = 0
    retiro: float = 0


HEADERS = ["prueba", "categoria", "clinica", "ingreso", "periodico", "retiro"]
# Ejemplos: Lima = precios sede Lima; TODAS = todas las sedes en provincia
EXAMPLE_ROWS = [
    ["Hemograma Completo", "Laboratorio", "Lima", 40, 35, 35],
    ["Marihuana cualitativo", "Laboratorio", "TODAS", 50, 50, 50],
]


def _build_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("No active sheet")
    ws.title = "Precios"
    for col, h in enumerate(HEADERS, 1):
        ws.cell(row=1, column=col, value=h)
    for row_idx, row_data in enumerate(EXAMPLE_ROWS, 2):
        for col, v in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=v)
    for col in range(1, len(HEADERS) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _parse_xlsx_rows(content: bytes) -> List[Dict[str, Any]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return []
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        return []
    # Normalizar nombres de columna (minúsculas, sin espacios)
    col_names = [str(h).strip().lower() if h is not None else "" for h in header]
    # Índices por nombre esperado
    idx = {}
    for name in ["prueba", "categoria", "clinica", "ingreso", "periodico", "retiro"]:
        try:
            idx[name] = col_names.index(name)
        except ValueError:
            pass
    if "prueba" not in idx or "categoria" not in idx:
        return []
    out = []
    for row in rows_iter:
        if row is None:
            continue
        row_list = list(row) if not isinstance(row, (list, tuple)) else row
        if len(row_list) <= max(idx.values()):
            continue
        prueba_val = row_list[idx["prueba"]] if idx["prueba"] < len(row_list) else None
        if prueba_val is None or (isinstance(prueba_val, str) and not prueba_val.strip()):
            continue
        def _cell(key: str, default: Any = "") -> Any:
            if key not in idx or idx[key] >= len(row_list):
                return default
            v = row_list[idx[key]]
            if v is None and default == 0:
                return 0
            if v is None:
                return ""
            return v

        out.append({
            "prueba": _cell("prueba", ""),
            "categoria": _cell("categoria", ""),
            "clinica": _cell("clinica", ""),
            "ingreso": _cell("ingreso", 0),
            "periodico": _cell("periodico", 0),
            "retiro": _cell("retiro", 0),
        })
    return out


@router.get("/template")
def download_template():
    """Descarga plantilla XLSX para importar precios."""
    if Workbook is None:
        raise HTTPException(status_code=500, detail="Error al generar la plantilla.")
    data = _build_template_xlsx()
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=plantilla_precios.xlsx"},
    )


@router.get("/search")
def search_tests(q: str = Query("", description="Búsqueda por nombre de prueba (mín. 2 caracteres)")):
    """Busca pruebas por nombre. Devuelve clínicas con precios y clínicas sin precio."""
    q_trim = (q or "").strip()
    if len(q_trim) < 2:
        return {"tests": []}
    db: Session = SessionLocal()
    try:
        tests = db.query(Test).filter(Test.name.ilike(f"%{q_trim}%")).order_by(Test.category, Test.name).limit(50).all()
        if not tests:
            return {"tests": []}
        test_ids = [t.id for t in tests]
        prices = db.query(Price).filter(Price.test_id.in_(test_ids)).all()
        clinic_ids = {p.clinic_id for p in prices if p.clinic_id is not None}
        clinics = {c.id: c.name for c in db.query(Clinic).filter(Clinic.id.in_(clinic_ids)).all()} if clinic_ids else {}
        all_clinic_names = ["Lima"] + [c.name for c in db.query(Clinic.name).order_by(Clinic.name).all()]
        prices_by_test = {}
        for p in prices:
            key = p.test_id
            if key not in prices_by_test:
                prices_by_test[key] = []
            clinic_name = "Lima" if p.clinic_id is None else clinics.get(p.clinic_id, "?")
            prices_by_test[key].append({
                "clinic_name": clinic_name,
                "clinic_id": p.clinic_id,
                "ingreso": float(p.ingreso),
                "periodico": float(p.periodico),
                "retiro": float(p.retiro),
            })
        result = []
        for t in tests:
            with_prices = prices_by_test.get(t.id, [])
            clinics_with_price = {row["clinic_name"] for row in with_prices}
            clinics_without_price = [c for c in all_clinic_names if c not in clinics_with_price]
            result.append({
                "test_id": t.id,
                "test_name": t.name,
                "category": t.category,
                "clinics_with_price": with_prices,
                "clinics_without_price": clinics_without_price,
            })
        return {"tests": result}
    finally:
        db.close()


@router.get("/list")
def list_prices_by_clinic(clinic: str = Query(..., description="Lima o nombre de la sede en provincia")):
    """Lista todos los exámenes con sus precios para la sede indicada (Lima o nombre de clínica)."""
    db: Session = SessionLocal()
    try:
        is_lima = clinic.strip().lower() in ("lima", "")
        clinic_id: Optional[int] = None
        if not is_lima:
            c = db.query(Clinic.id).filter(Clinic.name == clinic.strip()).first()
            if not c:
                raise HTTPException(status_code=404, detail="Sede no encontrada.")
            clinic_id = c.id
        tests = db.query(Test).order_by(Test.category, Test.name).all()
        if is_lima:
            prices_q = db.query(Price).filter(Price.clinic_id.is_(None))
        else:
            prices_q = db.query(Price).filter(Price.clinic_id == clinic_id)
        prices_by_test = {(p.test_id, p.clinic_id): p for p in prices_q.all()}
        rows = []
        for t in tests:
            p = prices_by_test.get((t.id, clinic_id))
            if p:
                rows.append(PriceRow(test_id=t.id, test_name=t.name, category=t.category, price_id=p.id, ingreso=p.ingreso, periodico=p.periodico, retiro=p.retiro))
            else:
                rows.append(PriceRow(test_id=t.id, test_name=t.name, category=t.category, price_id=None, ingreso=0, periodico=0, retiro=0))
        return {"clinic": "Lima" if is_lima else clinic.strip(), "clinic_id": clinic_id, "tests": [r.model_dump() for r in rows]}
    finally:
        db.close()


@router.post("/add")
def add_price(body: AddPriceBody):
    """Añade una prueba (crea test si no existe) y su precio para la sede. clinic_id null = Lima."""
    name = (body.test_name or "").strip()
    category = (body.category or "").strip()
    if not name or not category:
        raise HTTPException(status_code=400, detail="Nombre y categoría son obligatorios.")
    db: Session = SessionLocal()
    try:
        if body.clinic_id is not None:
            c = db.query(Clinic).filter(Clinic.id == body.clinic_id).first()
            if not c:
                raise HTTPException(status_code=404, detail="Sede no encontrada.")
        test = db.query(Test).filter(Test.name == name, Test.category == category).first()
        if not test:
            test = Test(name=name, category=category)
            db.add(test)
            db.flush()
        existing = (
            db.query(Price)
            .filter(Price.test_id == test.id, Price.clinic_id == body.clinic_id)
            .first()
        )
        if existing:
            existing.ingreso = max(0, float(body.ingreso))
            existing.periodico = max(0, float(body.periodico))
            existing.retiro = max(0, float(body.retiro))
            db.commit()
            db.refresh(existing)
            return {"id": existing.id, "test_id": existing.test_id, "test_name": test.name, "category": test.category, "clinic_id": existing.clinic_id, "ingreso": existing.ingreso, "periodico": existing.periodico, "retiro": existing.retiro}
        new_price = Price(
            test_id=test.id,
            clinic_id=body.clinic_id,
            ingreso=max(0, float(body.ingreso)),
            periodico=max(0, float(body.periodico)),
            retiro=max(0, float(body.retiro)),
        )
        db.add(new_price)
        db.commit()
        db.refresh(new_price)
        return {"id": new_price.id, "test_id": new_price.test_id, "test_name": test.name, "category": test.category, "clinic_id": new_price.clinic_id, "ingreso": new_price.ingreso, "periodico": new_price.periodico, "retiro": new_price.retiro}
    except HTTPException:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        raise HTTPException(status_code=500, detail="Error al guardar. Intenta de nuevo.")
    finally:
        db.close()


@router.put("")
def update_price(body: PriceUpdateBody):
    """Crea o actualiza un precio (test_id + clinic_id null = Lima)."""
    db: Session = SessionLocal()
    try:
        existing = (
            db.query(Price)
            .filter(Price.test_id == body.test_id, Price.clinic_id == body.clinic_id)
            .first()
        )
        if existing:
            existing.ingreso = max(0, float(body.ingreso))
            existing.periodico = max(0, float(body.periodico))
            existing.retiro = max(0, float(body.retiro))
            db.commit()
            return {"id": existing.id, "test_id": existing.test_id, "clinic_id": existing.clinic_id, "ingreso": existing.ingreso, "periodico": existing.periodico, "retiro": existing.retiro}
        test = db.query(Test).filter(Test.id == body.test_id).first()
        if not test:
            raise HTTPException(status_code=404, detail="Prueba no encontrada.")
        if body.clinic_id is not None:
            c = db.query(Clinic).filter(Clinic.id == body.clinic_id).first()
            if not c:
                raise HTTPException(status_code=404, detail="Sede no encontrada.")
        new_price = Price(
            test_id=body.test_id,
            clinic_id=body.clinic_id,
            ingreso=max(0, float(body.ingreso)),
            periodico=max(0, float(body.periodico)),
            retiro=max(0, float(body.retiro)),
        )
        db.add(new_price)
        db.commit()
        db.refresh(new_price)
        return {"id": new_price.id, "test_id": new_price.test_id, "clinic_id": new_price.clinic_id, "ingreso": new_price.ingreso, "periodico": new_price.periodico, "retiro": new_price.retiro}
    except HTTPException:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        raise HTTPException(status_code=500, detail="Error al guardar. Intenta de nuevo.")
    finally:
        db.close()


@router.post("/import")
def import_prices(file: UploadFile = File(...)):
    """Carga precios desde un archivo XLSX (mismo formato que la plantilla)."""
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Debes subir un archivo .xlsx")
    content = file.file.read()
    try:
        rows = _parse_xlsx_rows(content)
    except Exception:
        raise HTTPException(status_code=400, detail="El archivo no es válido.")
    if not rows:
        raise HTTPException(status_code=400, detail="El archivo no tiene filas de datos. Usa la plantilla con los encabezados indicados.")
    db: Session = SessionLocal()
    try:
        rows_done, errors = import_prices_from_rows(db, rows)
        db.commit()
        return {"imported": rows_done, "errors": errors}
    except Exception:
        db.rollback()
        raise HTTPException(status_code=500, detail="Error al importar. Intenta de nuevo.")
    finally:
        db.close()


@router.delete("")
def delete_price(body: DeletePriceBody):
    """
    Elimina precio(s) de una prueba según el alcance:
    - "clinic": elimina el precio para la clínica especificada (clinic_id requerido)
    - "lima": elimina el precio de Lima (clinic_id NULL)
    - "all_provincia": elimina todos los precios de provincia (todas las clínicas)
    - "all": elimina TODOS los precios (Lima + todas las clínicas) y la prueba si no tiene más datos
    """
    db: Session = SessionLocal()
    try:
        test = db.query(Test).filter(Test.id == body.test_id).first()
        if not test:
            raise HTTPException(status_code=404, detail="Prueba no encontrada.")
        
        deleted_count = 0
        if body.scope == "clinic":
            if body.clinic_id is None:
                raise HTTPException(status_code=400, detail="clinic_id es requerido para scope='clinic'.")
            deleted_count = db.query(Price).filter(
                Price.test_id == body.test_id,
                Price.clinic_id == body.clinic_id
            ).delete()
        elif body.scope == "lima":
            deleted_count = db.query(Price).filter(
                Price.test_id == body.test_id,
                Price.clinic_id.is_(None)
            ).delete()
        elif body.scope == "all_provincia":
            deleted_count = db.query(Price).filter(
                Price.test_id == body.test_id,
                Price.clinic_id.isnot(None)
            ).delete()
        elif body.scope == "all":
            deleted_count = db.query(Price).filter(
                Price.test_id == body.test_id
            ).delete()
            # Si no quedan precios para este test, eliminar el test también
            remaining = db.query(Price).filter(Price.test_id == body.test_id).count()
            if remaining == 0:
                db.delete(test)
        else:
            raise HTTPException(status_code=400, detail="scope inválido. Usa: clinic, lima, all_provincia, all.")
        
        db.commit()
        return {"deleted": deleted_count, "test_name": test.name}
    except HTTPException:
        db.rollback()
        raise
    except Exception:
        db.rollback()
        raise HTTPException(status_code=500, detail="Error al eliminar. Intenta de nuevo.")
    finally:
        db.close()
